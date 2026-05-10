from pathlib import Path
import json
from datetime import datetime
import openpyxl

PROJECT_ROOT = Path(r"C:\Projects\forest-harvest-system")
NFI4_ROOT = PROJECT_ROOT / "data" / "raw" / "nfi" / "nfi4"
DOCS_DATA = PROJECT_ROOT / "docs" / "data"
OBSIDIAN_NFI = Path(r"C:\ObsidianVaults\ForestHarvestSystem\database\nfi")

excel_files = [
    p for p in NFI4_ROOT.rglob("*.xlsx")
    if not p.name.startswith("~$")
]

result = {
    "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    "purpose": "Inspect NFI4 Excel workbooks for possible plot-level and tree-level tables.",
    "excel_file_count": len(excel_files),
    "workbooks": []
}

def clean(value):
    if value is None:
        return ""
    return str(value).strip()

for excel_path in excel_files:
    workbook_info = {
        "file": str(excel_path),
        "name": excel_path.name,
        "size_mb": round(excel_path.stat().st_size / 1024 / 1024, 3),
        "sheets": []
    }

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    for ws in wb.worksheets:
        sheet_info = {
            "sheet_name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "headers": [],
            "sample_rows": []
        }

        rows = ws.iter_rows(values_only=True)

        try:
            header_row = next(rows)
        except StopIteration:
            header_row = []

        headers = [clean(v) for v in header_row]
        sheet_info["headers"] = headers

        for i, row in enumerate(rows):
            if i >= 5:
                break

            row_values = [clean(v) for v in row]
            sheet_info["sample_rows"].append(row_values)

        workbook_info["sheets"].append(sheet_info)

    result["workbooks"].append(workbook_info)

DOCS_DATA.mkdir(parents=True, exist_ok=True)
OBSIDIAN_NFI.mkdir(parents=True, exist_ok=True)

json_path = DOCS_DATA / "nfi4-excel-structure-inventory.json"
md_path = DOCS_DATA / "nfi4-excel-structure-inventory.md"

json_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
(OBSIDIAN_NFI / "nfi4-excel-structure-inventory.json").write_text(
    json.dumps(result, ensure_ascii=False, indent=2),
    encoding="utf-8"
)

lines = []
lines.append("# NFI4 Excel 結構盤點")
lines.append("")
lines.append(f"產生時間：{result['generated_at']}")
lines.append("")
lines.append("## 一、目的")
lines.append("")
lines.append("本文件盤點 NFI4 Excel 檔案中的工作表、欄位名稱與前幾筆資料，用於判斷是否存在真正的樣木層級資料。")
lines.append("")
lines.append("## 二、Excel 檔案數")
lines.append("")
lines.append(str(result["excel_file_count"]))
lines.append("")

for workbook in result["workbooks"]:
    lines.append(f"## {workbook['name']}")
    lines.append("")
    lines.append(f"- 檔案路徑：`{workbook['file']}`")
    lines.append(f"- 大小 MB：{workbook['size_mb']}")
    lines.append("")

    for sheet in workbook["sheets"]:
        lines.append(f"### 工作表：{sheet['sheet_name']}")
        lines.append("")
        lines.append(f"- 列數：{sheet['max_row']}")
        lines.append(f"- 欄數：{sheet['max_column']}")
        lines.append("")
        lines.append("#### 欄位名稱")
        lines.append("")
        for h in sheet["headers"]:
            if h:
                lines.append(f"- {h}")
        lines.append("")
        lines.append("#### 前 5 筆資料")
        lines.append("")
        for row in sheet["sample_rows"]:
            preview = " | ".join(row[:20])
            lines.append(f"- {preview}")
        lines.append("")

lines.append("## 三、初步判讀方向")
lines.append("")
lines.append("若 Excel 中存在 DBH、胸徑、樹種、樣木號、樹高等欄位，則該 Excel 可能包含真正樣木資料。")
lines.append("")
lines.append("若 Excel 只包含 Height、Volumn、Crown、樣區BA、樣區Vol、CO2_ha 等欄位，則較可能仍為樣區或林分層級資料。")

md_text = "\n".join(lines)
md_path.write_text(md_text, encoding="utf-8")
(OBSIDIAN_NFI / "nfi4-excel-structure-inventory.md").write_text(md_text, encoding="utf-8")

print("NFI4 Excel structure inventory completed")
print("Excel file count:", result["excel_file_count"])
print()

for workbook in result["workbooks"]:
    print("Workbook:", workbook["name"], "size_mb=", workbook["size_mb"])
    for sheet in workbook["sheets"]:
        print("  Sheet:", sheet["sheet_name"], "rows=", sheet["max_row"], "cols=", sheet["max_column"])
        print("  Headers:")
        for h in sheet["headers"]:
            if h:
                print("   -", h)
        print()

print(json_path)
print(md_path)
